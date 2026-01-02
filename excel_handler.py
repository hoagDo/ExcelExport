# excel_handler.py
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class ExcelHandler:
    def __init__(self, template_path=None):
        self.template_path = template_path or "TrachNG_CN.xlsx"
        self.default_headers = [
            "Question Text",
            "Question Type",
            "Option 1",
            "Option 2",
            "Option 3",
            "Option 4",
            "Correct Answer",
            "Time in seconds",
            "Image Link",
            "Answer explanation"
        ]
    
    def create_template_if_not_exists(self):
        """Tạo template nếu không tồn tại"""
        if not os.path.exists(self.template_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Questions"
            
            # Ghi headers
            for col, header in enumerate(self.default_headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Đặt độ rộng cột
            column_widths = [60, 15, 40, 40, 40, 40, 15, 15, 30, 50]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            wb.save(self.template_path)
    
    def get_next_empty_row(self, ws):
        """Tìm dòng trống tiếp theo để ghi dữ liệu"""
        row = 3  # Bắt đầu từ dòng 3 (dòng 1-2 là header)
        while ws.cell(row=row, column=1).value is not None:
            row += 1
        return row
    
    def write_questions(self, questions, output_path=None):
        """Ghi câu hỏi vào file Excel"""
        output_path = output_path or self.template_path
        
        # Đảm bảo template tồn tại
        self.create_template_if_not_exists()
        
        # Mở workbook
        wb = load_workbook(output_path)
        if "Questions" not in wb.sheetnames:
            ws = wb.create_sheet("Questions")
        else:
            ws = wb["Questions"]
        
        # Tìm dòng bắt đầu
        start_row = self.get_next_empty_row(ws)
        
        # Ghi dữ liệu
        for i, question_data in enumerate(questions):
            row = start_row + i
            
            # Ghi câu hỏi
            ws.cell(row=row, column=1, value=question_data['question_text'])
            
            # Ghi loại câu hỏi
            ws.cell(row=row, column=2, value="Multiple Choice")
            
            # Ghi các lựa chọn
            options = question_data['options']
            for j in range(4):
                option_value = options[j] if j < len(options) else ""
                ws.cell(row=row, column=3 + j, value=option_value)
            
            # Định dạng hàng
            fill_color = "FFFFFF" if row % 2 == 0 else "F2F2F2"
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                # Đặt border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # Lưu file
        wb.save(output_path)
        
        return {
            'total_questions': len(questions),
            'start_row': start_row,
            'output_path': output_path
        }
    
    def export_summary(self, questions, stats):
        """Xuất file summary"""
        summary_wb = openpyxl.Workbook()
        ws = summary_wb.active
        ws.title = "Summary"
        
        # Ghi thông tin tổng quan
        ws.append(["Export Summary", ""])
        ws.append(["Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["Total Questions Processed", len(questions)])
        ws.append(["Written to Excel", stats.get('written', 0)])
        ws.append(["Skipped (duplicates)", stats.get('skipped', 0)])
        ws.append(["Merged", stats.get('merged', 0)])
        ws.append(["", ""])
        
        # Ghi danh sách câu hỏi
        ws.append(["Question List", ""])
        headers = ["No.", "Question Text", "Options Count"]
        ws.append(headers)
        
        for i, q in enumerate(questions, 1):
            ws.append([
                i,
                q['question_text'][:100] + "..." if len(q['question_text']) > 100 else q['question_text'],
                len(q['options'])
            ])
        
        # Định dạng
        for row in ws.iter_rows(min_row=1, max_row=7, max_col=2):
            for cell in row:
                cell.font = Font(bold=True)
        
        summary_path = "export_summary.xlsx"
        summary_wb.save(summary_path)
        
        return summary_path